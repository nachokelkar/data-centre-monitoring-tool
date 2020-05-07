import happybase
import os
import sys
import openpyxl as xl

conn = happybase.Connection('localhost', port=9090)


ips = dict()

# Input from Excel sheet
# wb is the workbook
# ws is the active sheet
# rows is a list of numbers used as a coordinate range
wb = xl.load_workbook(os.path.join(sys.path[0], '../input.xlsx'))
ws = wb.active
rows = list(map(str, range(2, ws.max_row+1)))

# Iterate through cells in col A
# For each cell, add IP to ips and set its value as an empty dict
for r in rows:
    ips[ws["A" +r].value] = dict()

wb.close()

conn.open()
conn.create_table('snmp', ips)
conn.create_table('ping', ips)
conn.create_table('ssh', ips)
conn.close()

conn.open()
print("TABLES CREATED: ", conn.tables())
conn.close()